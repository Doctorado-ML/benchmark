import os
from openpyxl import load_workbook
from xlsxwriter import Workbook
from .TestBase import TestBase
from ..ResultsFiles import Excel
from ..Utils import Folders


class ExcelTest(TestBase):
    def tearDown(self) -> None:
        files = [
            "results_accuracy_STree_iMac27_2021-09-30_11:42:07_0.xlsx",
            "results_accuracy_STree_iMac27_2021-10-27_09:40:40_0.xlsx",
            "results_accuracy_ODTE_Galgo_2022-04-20_10:52:20_0.xlsx",
        ]
        self.remove_files(files, Folders.excel)
        return super().tearDown()

    def test_report_excel_compared(self):
        file_name = "results_accuracy_STree_iMac27_2021-09-30_11:42:07_0.json"
        report = Excel(file_name, compare=True)
        report.report()
        file_output = report.get_file_name()
        book = load_workbook(os.path.join(Folders.excel, file_output))
        sheet = book["STree"]
        self.check_excel_sheet(sheet, "excel_compared")

    def test_report_excel(self):
        file_name = "results_accuracy_STree_iMac27_2021-09-30_11:42:07_0.json"
        report = Excel(file_name, compare=False)
        report.report()
        file_output = report.get_file_name()
        book = load_workbook(os.path.join(Folders.excel, file_output))
        sheet = book["STree"]
        self.check_excel_sheet(sheet, "excel")

    def test_Excel_Add_sheet(self):
        file_name = "results_accuracy_STree_iMac27_2021-10-27_09:40:40_0.json"
        excel_file_name = file_name.replace(".json", ".xlsx")
        book = Workbook(os.path.join(Folders.excel, excel_file_name))
        excel = Excel(file_name=file_name, book=book)
        excel.report()
        report = Excel(
            file_name="results_accuracy_ODTE_Galgo_2022-04-20_10:52:20_0.json",
            book=book,
        )
        report.report()
        book.close()
        book = load_workbook(os.path.join(Folders.excel, excel_file_name))
        sheet = book["STree"]
        self.check_excel_sheet(sheet, "excel_add_STree")
        sheet = book["ODTE"]
        self.check_excel_sheet(sheet, "excel_add_ODTE")
