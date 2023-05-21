import os
import shutil
from unittest.mock import patch
from openpyxl import load_workbook
from ...Utils import Folders, Files, NO_RESULTS
from ..TestBase import TestBase


class BeListTest(TestBase):
    def setUp(self):
        self.prepare_scripts_env()

    @patch("benchmark.Results.get_input", return_value="q")
    def test_be_list(self, input_data):
        stdout, stderr = self.execute_script("be_list", ["-m", "STree"])
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "be_list_model")

    @patch("benchmark.Results.get_input", side_effect=iter(["x", "q"]))
    def test_be_list_invalid_option(self, input_data):
        stdout, stderr = self.execute_script("be_list", ["-m", "STree"])
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "be_list_model_invalid")

    @patch("benchmark.Results.get_input", side_effect=iter(["0", "q"]))
    def test_be_list_report(self, input_data):
        stdout, stderr = self.execute_script("be_list", ["-m", "STree"])
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "be_list_report")

    @patch("benchmark.Results.get_input", side_effect=iter(["r", "q"]))
    def test_be_list_twice(self, input_data):
        stdout, stderr = self.execute_script("be_list", ["-m", "STree"])
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "be_list_model_2")

    @patch("benchmark.Results.get_input", side_effect=iter(["e 2", "q"]))
    def test_be_list_report_excel(self, input_data):
        stdout, stderr = self.execute_script("be_list", ["-m", "STree"])
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "be_list_report_excel")

        book = load_workbook(os.path.join(Folders.excel, Files.be_list_excel))
        sheet = book["STree"]
        self.check_excel_sheet(sheet, "excel")

    @patch(
        "benchmark.Results.get_input",
        side_effect=iter(["e 2", "e 1", "q"]),
    )
    def test_be_list_report_excel_twice(self, input_data):
        stdout, stderr = self.execute_script("be_list", ["-m", "STree"])
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "be_list_report_excel_2")
        book = load_workbook(os.path.join(Folders.excel, Files.be_list_excel))
        sheet = book["STree"]
        self.check_excel_sheet(sheet, "excel")
        sheet = book["STree2"]
        self.check_excel_sheet(sheet, "excel2")

    @patch("benchmark.Results.get_input", return_value="q")
    def test_be_list_no_data(self, input_data):
        stdout, stderr = self.execute_script(
            "be_list", ["-m", "Wodt", "-s", "f1-macro"]
        )
        self.assertEqual(stderr.getvalue(), "")
        self.assertEqual(stdout.getvalue(), f"{NO_RESULTS}\n")

    @patch(
        "benchmark.Results.get_input",
        side_effect=iter(["d 0", "y", "", "q"]),
    )
    # @patch("benchmark.ResultsBase.get_input", side_effect=iter(["q"]))
    def test_be_list_delete(self, input_data):
        def copy_files(source_folder, target_folder, file_name):
            source = os.path.join(source_folder, file_name)
            target = os.path.join(target_folder, file_name)
            shutil.copyfile(source, target)

        file_name = (
            "results_accuracy_XGBoost_MacBookpro16_2022-05-04_11:00:"
            "35_0.json"
        )
        # move nan result from hidden to results
        copy_files(Folders.hidden_results, Folders.results, file_name)
        try:
            # list and delete result
            stdout, stderr = self.execute_script("be_list", "")
            self.assertEqual(stderr.getvalue(), "")
            self.check_output_file(stdout, "be_list_delete")
        except Exception:
            # delete the result copied if be_list couldn't
            os.unlink(os.path.join(Folders.results, file_name))
            self.fail("test_be_list_delete() should not raise exception")

    @patch(
        "benchmark.Results.get_input",
        side_effect=iter(["h 0", "y", "", "q"]),
    )
    def test_be_list_hide(self, input_data):
        def swap_files(source_folder, target_folder, file_name):
            source = os.path.join(source_folder, file_name)
            target = os.path.join(target_folder, file_name)
            os.rename(source, target)

        file_name = (
            "results_accuracy_XGBoost_MacBookpro16_2022-05-04_11:00:"
            "35_0.json"
        )
        # move nan result from hidden to results
        swap_files(Folders.hidden_results, Folders.results, file_name)
        try:
            # list and move nan result to hidden again
            stdout, stderr = self.execute_script("be_list", "")
            self.assertEqual(stderr.getvalue(), "")
            self.check_output_file(stdout, "be_list_hide")
        except Exception:
            # delete the result copied if be_list couldn't
            swap_files(Folders.results, Folders.hidden_results, file_name)
            self.fail("test_be_list_hide() should not raise exception")

    @patch("benchmark.Results.get_input", side_effect=iter(["h 0", "q"]))
    def test_be_list_already_hidden(self, input_data):
        stdout, stderr = self.execute_script("be_list", ["--hidden"])
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "be_list_already_hidden")

    @patch("benchmark.Results.get_input", side_effect=iter(["h 0", "n", "q"]))
    def test_be_list_dont_hide(self, input_data):
        stdout, stderr = self.execute_script("be_list", "")
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "be_list_default")

    @patch("benchmark.Results.get_input", side_effect=iter(["q"]))
    def test_be_list_hidden_nan(self, input_data):
        stdout, stderr = self.execute_script("be_list", ["--hidden", "--nan"])
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "be_list_hidden_nan")

    @patch("benchmark.Results.get_input", side_effect=iter(["q"]))
    def test_be_list_hidden(self, input_data):
        stdout, stderr = self.execute_script("be_list", ["--hidden"])
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "be_list_hidden")

    def test_be_no_env(self):
        path = os.getcwd()
        os.chdir("..")
        stderr = None
        try:
            _, stderr = self.execute_script("be_list", [])
        except SystemExit as e:
            self.assertEqual(e.code, 1)
        finally:
            os.chdir(path)
        self.assertIsNone(stderr)
