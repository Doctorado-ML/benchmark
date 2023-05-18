import os
from openpyxl import load_workbook
from io import StringIO
from unittest.mock import patch
from ...Utils import Folders, Files
from ..TestBase import TestBase
from ..._version import __version__


class BeReportTest(TestBase):
    def setUp(self):
        self.prepare_scripts_env()

    def tearDown(self) -> None:
        files = [
            "results_accuracy_ODTE_Galgo_2022-04-20_10:52:20_0.sql",
            "results_accuracy_STree_iMac27_2021-09-30_11:42:07_0.xlsx",
        ]
        self.remove_files(files, Folders.results)
        self.remove_files(
            [Files.datasets_report_excel],
            os.path.join(os.getcwd(), Folders.excel),
        )
        return super().tearDown()

    def test_be_report(self):
        file_name = os.path.join(
            "results",
            "results_accuracy_STree_iMac27_2021-09-30_11:42:07_0.json",
        )
        stdout, stderr = self.execute_script("be_report", ["file", file_name])
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "report")

    def test_be_report_not_found(self):
        stdout, stderr = self.execute_script("be_report", ["file", "unknown"])
        self.assertEqual(stderr.getvalue(), "")
        self.assertEqual(stdout.getvalue(), "unknown does not exists!\n")

    def test_be_report_compare(self):
        file_name = "results_accuracy_STree_iMac27_2021-09-30_11:42:07_0.json"
        stdout, stderr = self.execute_script(
            "be_report", ["file", file_name, "-c"]
        )
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "report_compared")

    def test_be_report_datatsets(self):
        stdout, stderr = self.execute_script("be_report", ["datasets"])
        self.assertEqual(stderr.getvalue(), "")
        file_name = f"report_datasets{self.ext}"
        with open(os.path.join(self.test_files, file_name)) as f:
            expected = f.read()
        output_text = stdout.getvalue().splitlines()
        for line, index in zip(expected.splitlines(), range(len(expected))):
            if self.benchmark_version in line:
                # replace benchmark version
                line = self.replace_benchmark_version(line, output_text, index)
            self.assertEqual(line, output_text[index])

    def test_be_report_datasets_excel(self):
        stdout, stderr = self.execute_script("be_report", ["datasets", "-x"])
        self.assertEqual(stderr.getvalue(), "")
        file_name = f"report_datasets{self.ext}"
        with open(os.path.join(self.test_files, file_name)) as f:
            expected = f.read()
        output_text = stdout.getvalue().splitlines()
        for line, index in zip(expected.splitlines(), range(len(expected))):
            if self.benchmark_version in line:
                # replace benchmark version
                line = self.replace_benchmark_version(line, output_text, index)
            self.assertEqual(line, output_text[index])
        file_name = os.path.join(
            os.getcwd(), Folders.excel, Files.datasets_report_excel
        )
        book = load_workbook(file_name)
        sheet = book["Datasets"]
        self.check_excel_sheet(
            sheet,
            "exreport_excel_Datasets",
            replace=self.benchmark_version,
            with_this=__version__,
        )

    def test_be_report_best(self):
        stdout, stderr = self.execute_script(
            "be_report", ["best", "-s", "accuracy", "-m", "STree"]
        )
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "report_best")

    def test_be_report_grid(self):
        stdout, stderr = self.execute_script(
            "be_report", ["grid", "-s", "accuracy", "-m", "STree"]
        )
        self.assertEqual(stderr.getvalue(), "")
        file_name = "report_grid.test"
        with open(os.path.join(self.test_files, file_name)) as f:
            expected = f.read().splitlines()
        output_text = stdout.getvalue().splitlines()
        # Compare replacing STree version
        for line, index in zip(expected, range(len(expected))):
            if "1.2.4" in line:
                # replace STree version
                line = self.replace_STree_version(line, output_text, index)
            self.assertEqual(line, output_text[index])

    @patch("sys.stderr", new_callable=StringIO)
    def test_be_report_unknown_subcommand(self, stderr):
        with self.assertRaises(SystemExit) as msg:
            module = self.search_script("be_report")
            module.main(["unknown"])
        self.assertEqual(msg.exception.code, 2)
        self.check_output_file(stderr, "report_unknown_subcommand")

    def test_be_report_without_subcommand(self):
        stdout, stderr = self.execute_script("be_report", "")
        self.assertEqual(stderr.getvalue(), "")
        self.maxDiff = None
        # Can't use check_output_file because of the width of the console
        # output is different in different environments
        file_name = "report_without_subcommand" + self.ext
        with open(os.path.join(self.test_files, file_name)) as f:
            expected = f.read()
        if expected == stdout.getvalue():
            self.assertEqual(stdout.getvalue(), expected)
        else:
            self.check_output_file(stdout, "report_without_subcommand2")

    def test_be_report_excel_compared(self):
        file_name = "results_accuracy_STree_iMac27_2021-09-30_11:42:07_0.json"
        stdout, stderr = self.execute_script(
            "be_report",
            ["file", file_name, "-x", "-c"],
        )
        file_name = os.path.join(
            Folders.excel, file_name.replace(Files.report_ext, ".xlsx")
        )
        book = load_workbook(file_name)
        sheet = book["STree"]
        self.check_excel_sheet(sheet, "excel_compared")
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "report_compared")

    def test_be_report_excel(self):
        file_name = "results_accuracy_STree_iMac27_2021-09-30_11:42:07_0.json"
        stdout, stderr = self.execute_script(
            "be_report",
            ["file", file_name, "-x"],
        )
        file_name = os.path.join(
            Folders.excel, file_name.replace(".json", ".xlsx")
        )
        book = load_workbook(file_name)
        sheet = book["STree"]
        self.check_excel_sheet(sheet, "excel")
        self.assertEqual(stderr.getvalue(), "")
        self.check_output_file(stdout, "report")

    def test_be_report_sql(self):
        file_name = "results_accuracy_ODTE_Galgo_2022-04-20_10:52:20_0.json"
        stdout, stderr = self.execute_script(
            "be_report",
            ["file", file_name, "-q"],
        )
        file_name = os.path.join(
            Folders.results, file_name.replace(".json", ".sql")
        )
        self.check_file_file(file_name, "sql")
        self.assertEqual(stderr.getvalue(), "")
