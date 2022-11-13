import os
from openpyxl import load_workbook
from ...Utils import NO_RESULTS, Folders, Files
from ..TestBase import TestBase
from ..._version import __version__


class BeBenchmarkTest(TestBase):
    def setUp(self):
        self.prepare_scripts_env()
        self.score = "accuracy"

    def tearDown(self) -> None:
        files = []
        for score in [self.score, "unknown"]:
            files.append(Files.exreport(score))
            files.append(Files.exreport_output(score))
            files.append(Files.exreport_err(score))
        files.append(Files.exreport_excel(self.score))
        files.append(Files.exreport_pdf)
        files.append(Files.tex_output(self.score))
        self.remove_files(files, Folders.exreport)
        self.remove_files(files, ".")
        return super().tearDown()

    def test_be_benchmark_complete(self):
        stdout, stderr = self.execute_script(
            "be_benchmark", ["-s", self.score, "-q", "1", "-t", "1", "-x", "1"]
        )
        self.assertEqual(stderr.getvalue(), "")
        # Check output
        self.check_output_file(stdout, "be_benchmark_complete")
        # Check csv file
        file_name = os.path.join(Folders.exreport, Files.exreport(self.score))
        self.check_file_file(file_name, "exreport_csv")
        # Check tex file
        file_name = os.path.join(
            Folders.exreport, Files.tex_output(self.score)
        )
        self.assertTrue(os.path.exists(file_name))
        self.check_file_file(file_name, "exreport_tex")
        # Check excel file
        file_name = os.path.join(
            Folders.exreport, Files.exreport_excel(self.score)
        )
        book = load_workbook(file_name)
        replace = None
        with_this = None
        for sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            if sheet_name == "Datasets":
                replace = self.benchmark_version
                with_this = __version__
            self.check_excel_sheet(
                sheet,
                f"exreport_excel_{sheet_name}",
                replace=replace,
                with_this=with_this,
            )

    def test_be_benchmark_single(self):
        stdout, stderr = self.execute_script(
            "be_benchmark", ["-s", self.score, "-q", "1"]
        )
        self.assertEqual(stderr.getvalue(), "")
        # Check output
        self.check_output_file(stdout, "be_benchmark")
        # Check csv file
        file_name = os.path.join(Folders.exreport, Files.exreport(self.score))
        self.check_file_file(file_name, "exreport_csv")

    def test_be_benchmark_no_data(self):
        stdout, stderr = self.execute_script(
            "be_benchmark", ["-s", "f1-weighted"]
        )
        self.assertEqual(stderr.getvalue(), "")
        self.assertEqual(stdout.getvalue(), f"{NO_RESULTS}\n")
