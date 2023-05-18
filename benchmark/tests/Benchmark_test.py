import os
from io import StringIO
from unittest.mock import patch
from openpyxl import load_workbook
from .TestBase import TestBase
from ..Utils import Folders, Files, NO_RESULTS
from ..Results import Benchmark
from .._version import __version__


class BenchmarkTest(TestBase):
    def tearDown(self) -> None:
        files = []
        for score in ["accuracy", "unknown"]:
            files.append(Files.exreport(score))
            files.append(Files.exreport_output(score))
            files.append(Files.exreport_err(score))
        files.append(Files.exreport_pdf)
        files.append(Files.tex_output("accuracy"))
        self.remove_files(files, Folders.exreport)
        self.remove_files([Files.exreport_excel("accuracy")], Folders.excel)
        self.remove_files(files, ".")
        return super().tearDown()

    def test_csv(self):
        benchmark = Benchmark("accuracy", visualize=False)
        benchmark.compile_results()
        benchmark.save_results()
        self.check_file_file(benchmark.get_result_file_name(), "exreport_csv")

    def test_exreport_report(self):
        benchmark = Benchmark("accuracy", visualize=False)
        benchmark.compile_results()
        benchmark.save_results()
        with patch(self.output, new=StringIO()) as stdout:
            benchmark.report(tex_output=False)
        self.check_output_file(stdout, "exreport_report")

    def test_exreport(self):
        benchmark = Benchmark("accuracy", visualize=False)
        benchmark.compile_results()
        benchmark.save_results()
        with patch(self.output, new=StringIO()) as stdout:
            benchmark.exreport()
        with open(os.path.join(self.test_files, "exreport.test")) as f:
            expected_t = f.read()
        computed_t = stdout.getvalue()
        computed_t = computed_t.split("\n")
        computed_t.pop(0)
        for computed, expected in zip(computed_t, expected_t.split("\n")):
            self.assertEqual(computed, expected)

    def test_exreport_remove_previous(self):
        os.makedirs(Folders.report)
        with open(os.path.join(Files.exreport_pdf), "w") as f:
            print("x", file=f)
        self.assertTrue(os.path.exists(Files.exreport_pdf))
        self.assertTrue(os.path.exists(Folders.report))
        benchmark = Benchmark("accuracy", visualize=False)
        benchmark.compile_results()
        benchmark.save_results()
        with patch(self.output, new=StringIO()):
            benchmark.exreport()
        self.assertFalse(os.path.exists(Files.exreport_pdf))
        self.assertFalse(os.path.exists(Folders.report))

    def test_exreport_error(self):
        benchmark = Benchmark("accuracy", visualize=False)
        benchmark.compile_results()
        benchmark.save_results()
        # Make Rscript exreport fail
        benchmark._score = "unknown"
        with patch(self.output, new=StringIO()) as stdout:
            benchmark.exreport()
        self.check_output_file(stdout, "exreport_error")

    def test_exreport_no_data(self):
        benchmark = Benchmark("f1-weighted", visualize=False)
        with self.assertRaises(ValueError) as msg:
            benchmark.compile_results()
        self.assertEqual(str(msg.exception), NO_RESULTS)

    def test_tex_output(self):
        benchmark = Benchmark("accuracy", visualize=False)
        benchmark.compile_results()
        benchmark.save_results()
        with patch(self.output, new=StringIO()) as stdout:
            benchmark.report(tex_output=True)
        self.check_output_file(stdout, "exreport_report")
        self.assertTrue(os.path.exists(benchmark.get_tex_file()))
        self.check_file_file(benchmark.get_tex_file(), "exreport_tex")

    @staticmethod
    def generate_excel_sheet(test, sheet, file_name):
        with open(os.path.join("test_files", file_name), "w") as f:
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    value = sheet.cell(row=row, column=col).value
                    if value is not None:
                        print(f'{row};{col};"{value}"', file=f)

    def test_excel_output(self):
        benchmark = Benchmark("accuracy", visualize=False)
        benchmark.compile_results()
        benchmark.save_results()
        with patch(self.output, new=StringIO()):
            benchmark.exreport()
        benchmark.excel()
        file_name = benchmark.get_excel_file_name()
        book = load_workbook(file_name)
        replace = None
        with_this = None
        for sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            if sheet_name == "Datasets":
                replace = self.benchmark_version
                with_this = __version__
            self.check_excel_sheet(
                sheet,
                f"exreport_excel_{sheet_name}",
                replace=replace,
                with_this=with_this,
            )
